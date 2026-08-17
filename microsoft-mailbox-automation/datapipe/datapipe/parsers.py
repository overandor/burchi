"""PDF and spreadsheet parsers that produce rows and entities."""

from __future__ import annotations

import csv
import os
from typing import Any

import openpyxl

from .extractors import (
    Entity,
    extract_entities_from_row,
    extract_entities_from_text,
)


# ---------------------------------------------------------------------------
# Spreadsheet parsing
# ---------------------------------------------------------------------------


def _score_sheet(ws) -> int:
    """Score a worksheet by data richness — higher is more likely the real data sheet."""
    score = 0
    # Count non-empty data rows (skip first 5 rows as potential headers)
    data_rows = 0
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 500), values_only=True):
        non_empty = sum(1 for c in row if c is not None and str(c).strip())
        if non_empty >= 2:
            data_rows += 1
    score += data_rows * 10

    # Bonus for having many columns
    if ws.max_column >= 5:
        score += 50
    if ws.max_column >= 10:
        score += 50

    # Bonus for entity-rich headers (names, addresses, doctors, etc.)
    header_text = ""
    for row in ws.iter_rows(min_row=1, max_row=min(20, ws.max_row), values_only=True):
        non_empty = sum(1 for c in row if c is not None and str(c).strip())
        if non_empty >= 3:
            header_text = " ".join(str(c).lower() for c in row if c is not None and str(c).strip())
            break

    entity_keywords = [
        "name", "doctor", "physician", "provider", "address", "street",
        "city", "state", "zip", "npi", "specialty", "practice", "clinic",
        "hospital", "phone", "email", "contact", "prescriber",
    ]
    entity_hits = sum(1 for kw in entity_keywords if kw in header_text)
    score += entity_hits * 200

    # Penalty for summary/total/pivot/dashboard sheets
    name_lower = ws.title.lower()
    if any(kw in name_lower for kw in ("summary", "total", "pivot", "dashboard", "overview")):
        score = score // 3

    # Penalty for very wide sheets (likely raw data dumps, not curated entity sheets)
    if ws.max_column > 100:
        score = score // 2

    # Penalty for data dump sheets (MASTER_DATA, GRAPH, etc.)
    if any(kw in name_lower for kw in ("master_data", "graph", "stat_", "terr_", "prod_", "version", "macros", "help")):
        score = score // 2

    return max(score, 0)


def _auto_detect_best_sheet(wb) -> str | None:
    """Pick the most data-rich sheet from a workbook."""
    scored = [(name, _score_sheet(wb[name])) for name in wb.sheetnames]
    scored.sort(key=lambda x: x[1], reverse=True)
    if not scored or scored[0][1] == 0:
        return None
    return scored[0][0]


def parse_xlsx(
    path: str,
    sheet_filter: str | None = None,
    auto_detect: bool = True,
) -> list[dict[str, Any]]:
    """Parse an .xlsx file and return a list of row dicts.

    Each row dict maps column header -> value.  Rows from all sheets are
    combined.  Sheet name is stored under ``_sheet``.
    If *sheet_filter* is given, only that sheet is parsed.
    If *auto_detect* is True and no *sheet_filter* is given, the most
    data-rich sheet is selected automatically (summary/total sheets are
    deprioritized).
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    rows: list[dict[str, Any]] = []

    sheets = wb.sheetnames
    if sheet_filter:
        sheets = [s for s in sheets if sheet_filter.lower() in s.lower()]
    elif auto_detect and len(sheets) > 1:
        best = _auto_detect_best_sheet(wb)
        if best:
            sheets = [best]

    for sheet_name in sheets:
        ws = wb[sheet_name]

        # Find the header row — first row where >= 3 cells are non-empty
        header_row_idx = None
        headers: list[str] = []
        for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(20, ws.max_row), values_only=True), start=1):
            non_empty = sum(1 for c in row if c is not None and str(c).strip())
            if non_empty >= 3:
                header_row_idx = r_idx
                headers = [str(c).strip() if c is not None else f"col_{i}" for i, c in enumerate(row)]
                break

        if header_row_idx is None:
            # No header found — treat every row as data with generic column names
            headers = [f"col_{i}" for i in range(ws.max_column)]
            header_row_idx = 0

        for r_idx, row in enumerate(
            ws.iter_rows(min_row=header_row_idx + 1, max_row=ws.max_row, values_only=True),
            start=header_row_idx + 1,
        ):
            if all(c is None or str(c).strip() == "" for c in row):
                continue
            row_dict: dict[str, Any] = {"_sheet": sheet_name, "_row": r_idx}
            for i, val in enumerate(row):
                if i < len(headers):
                    row_dict[headers[i]] = val
            rows.append(row_dict)

    return rows


def parse_csv(path: str) -> list[dict[str, Any]]:
    """Parse a CSV file and return a list of row dicts."""
    rows: list[dict[str, Any]] = []
    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for r_idx, row in enumerate(reader, start=2):
            row["_sheet"] = "csv"
            row["_row"] = r_idx
            rows.append(row)
    return rows


def parse_spreadsheet(
    path: str,
    sheet_filter: str | None = None,
    auto_detect: bool = True,
) -> list[dict[str, Any]]:
    """Dispatch to the correct spreadsheet parser based on extension."""
    ext = os.path.splitext(path)[1].lower()
    if ext == ".xlsx":
        return parse_xlsx(path, sheet_filter=sheet_filter, auto_detect=auto_detect)
    elif ext == ".csv":
        return parse_csv(path)
    elif ext == ".tsv":
        rows: list[dict[str, Any]] = []
        with open(path, newline="", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f, delimiter="\t")
            for r_idx, row in enumerate(reader, start=2):
                row["_sheet"] = "tsv"
                row["_row"] = r_idx
                rows.append(row)
        return rows
    else:
        raise ValueError(f"Unsupported spreadsheet format: {ext}")


# ---------------------------------------------------------------------------
# PDF parsing
# ---------------------------------------------------------------------------


def parse_pdf(path: str) -> list[dict[str, Any]]:
    """Parse a PDF and return a list of "row" dicts (one per page).

    Each row dict has keys: ``_sheet`` (always "pdf"), ``_page``,
    ``_row``, and ``text``.
    """
    import pdfplumber

    rows: list[dict[str, Any]] = []
    with pdfplumber.open(path) as pdf:
        for page_num, page in enumerate(pdf.pages, start=1):
            text = page.extract_text() or ""
            # Also extract tables if present
            tables = page.extract_tables()
            if tables:
                for t_idx, table in enumerate(tables):
                    if not table:
                        continue
                    headers = [str(c).strip() if c else f"col_{i}" for i, c in enumerate(table[0])]
                    for r_idx, row in enumerate(table[1:], start=2):
                        if all(c is None or str(c).strip() == "" for c in row):
                            continue
                        row_dict: dict[str, Any] = {
                            "_sheet": f"pdf_table_{t_idx}",
                            "_page": page_num,
                            "_row": r_idx,
                        }
                        for i, val in enumerate(row):
                            if i < len(headers):
                                row_dict[headers[i]] = val
                        rows.append(row_dict)

            # Always add the full-page text as a row
            if text.strip():
                rows.append({
                    "_sheet": "pdf",
                    "_page": page_num,
                    "_row": 1,
                    "text": text,
                })
    return rows


# ---------------------------------------------------------------------------
# Unified parser
# ---------------------------------------------------------------------------


def parse_file(
    path: str,
    sheet_filter: str | None = None,
    auto_detect: bool = True,
) -> list[dict[str, Any]]:
    """Parse any supported file type and return rows."""
    ext = os.path.splitext(path)[1].lower()
    if ext == ".pdf":
        return parse_pdf(path)
    elif ext in (".xlsx", ".csv", ".tsv"):
        return parse_spreadsheet(path, sheet_filter=sheet_filter, auto_detect=auto_detect)
    else:
        raise ValueError(f"Unsupported file type: {ext}")


def extract_entities_from_file(
    path: str,
    sheet_filter: str | None = None,
    auto_detect: bool = True,
) -> list[Entity]:
    """Parse a file and extract all entities from it."""
    rows = parse_file(path, sheet_filter=sheet_filter, auto_detect=auto_detect)
    entities: list[Entity] = []

    for row in rows:
        sheet = row.get("_sheet", "")
        row_num = row.get("_row")
        page = row.get("_page")

        if "text" in row and len(row) <= 4:
            # Free-text row (PDF page text)
            entities.extend(
                extract_entities_from_text(
                    row["text"], source_file=path, source_sheet=sheet,
                    source_row=row_num, source_page=page,
                )
            )
        else:
            # Structured row
            clean_row = {k: v for k, v in row.items() if k and not k.startswith("_")}
            entities.extend(
                extract_entities_from_row(
                    clean_row, source_file=path, source_sheet=sheet,
                    source_row=row_num,
                )
            )

    return entities
